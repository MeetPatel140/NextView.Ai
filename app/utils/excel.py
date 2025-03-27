from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl.workbook import Workbook

def create_workbook_with_default_style():
    """Create a new workbook with default style to prevent openpyxl warning."""
    wb = Workbook()
    
    # Create a default style
    default_style = NamedStyle(name='default')
    default_style.font = Font(name='Arial', size=11)
    default_style.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    default_style.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Add the style to workbook
    if 'default' not in wb.style_names:
        wb.add_named_style(default_style)
    
    return wb